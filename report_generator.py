from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak


def _safe(value):
    if value is None:
        return "-"
    return str(value)


def _nivel_color(nivel):
    if nivel == "Alto":
        return "FFCDD2"
    if nivel == "Medio":
        return "FFF9C4"
    if nivel == "Bajo":
        return "C8E6C9"
    return "E0E0E0"


def _ajustar_columnas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


def crear_excel_reporte(nombre, codigo, carrera, ciclo, diagnostico, diagnostico_detallado, resumen_tareas, cursos, tareas):
    output = BytesIO()
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Reporte AURA"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    datos = [
        ["Estudiante", nombre],
        ["Código", codigo],
        ["Carrera", carrera],
        ["Ciclo", ciclo],
    ]
    if diagnostico:
        horas, promedio, _, estres, motivacion, procrast, puntaje, riesgo, fecha = diagnostico
        datos.extend([
            ["Horas de estudio por día", horas],
            ["Promedio ponderado", promedio],
            ["Puntaje de riesgo", f"{puntaje}/100"],
            ["Nivel de riesgo", riesgo],
            ["Estrés", estres],
            ["Motivación", motivacion],
            ["Procrastinación", procrast],
            ["Fecha diagnóstico", fecha],
        ])
        if diagnostico_detallado:
            datos.extend([
                ["Estado de ánimo", diagnostico_detallado.get("indice_estado_animo")],
                ["Alerta emocional", "Sí" if diagnostico_detallado.get("alerta_emocional") == 1 else "No"],
                ["Diagnóstico IA", diagnostico_detallado.get("diagnostico_general_ia")],
                ["Recomendación estudiante", diagnostico_detallado.get("recomendacion_estudiante_ia")],
                ["Recomendación tutoría", diagnostico_detallado.get("recomendacion_tutoria_ia")],
            ])
    else:
        datos.append(["Diagnóstico", "Sin diagnóstico registrado"])

    datos.extend([
        ["Total de tareas", resumen_tareas.get("total", 0)],
        ["Tareas completadas", resumen_tareas.get("completadas", 0)],
        ["Tareas pendientes", resumen_tareas.get("pendientes", 0)],
        ["Tareas alta prioridad", resumen_tareas.get("alta_prioridad", 0)],
        ["Cumplimiento", f"{resumen_tareas.get('porcentaje_cumplimiento', 0)}%"],
    ])

    fila = 4
    for clave, valor in datos:
        ws.cell(row=fila, column=1, value=clave).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=valor)
        ws.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1
    _ajustar_columnas(ws)

    ws_cursos = wb.create_sheet("Cursos")
    ws_cursos.append(["ID", "Curso", "Docente", "Créditos", "Dificultad", "Estado"])
    for c in cursos:
        ws_cursos.append(list(c))
    for cell in ws_cursos[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    _ajustar_columnas(ws_cursos)

    ws_tareas = wb.create_sheet("Tareas")
    ws_tareas.append(["ID", "Tipo", "Actividad", "Curso", "Fecha de entrega", "Prioridad", "Estado"])
    for t in tareas:
        ws_tareas.append([_safe(x) for x in t])
    for cell in ws_tareas[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    _ajustar_columnas(ws_tareas)

    if diagnostico_detallado and diagnostico_detallado.get("respuestas"):
        ws_diag = wb.create_sheet("Encuesta")
        ws_diag.append(["Pregunta", "Respuesta"])
        for i, v in diagnostico_detallado["respuestas"].items():
            ws_diag.append([i, v])
        for cell in ws_diag[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        _ajustar_columnas(ws_diag)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _paragraph(text, style):
    return Paragraph(_safe(text).replace("\n", "<br/>").replace("&", "&amp;"), style)


def _tabla_pares(pares, style_texto):
    data = [[_paragraph(k, style_texto), _paragraph(v, style_texto)] for k, v in pares]
    tabla = Table(data, colWidths=[5.0 * cm, 11.0 * cm], hAlign="LEFT")
    tabla.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EEF7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tabla


def crear_pdf_reporte(nombre, codigo, carrera, ciclo, diagnostico, diagnostico_detallado, resumen_tareas, cursos, tareas):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.6 * cm,
        leftMargin=1.6 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle("AuraTitle", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=12)
    h2 = ParagraphStyle("AuraH2", parent=styles["Heading2"], fontSize=13, leading=16, spaceBefore=10, spaceAfter=6)
    normal = ParagraphStyle("AuraNormal", parent=styles["BodyText"], fontSize=9.5, leading=13, spaceAfter=4, alignment=TA_JUSTIFY)
    small = ParagraphStyle("AuraSmall", parent=styles["BodyText"], fontSize=8.5, leading=11, alignment=TA_LEFT)

    elems = []
    elems.append(Paragraph("Reporte AURA", title))
    elems.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal))
    elems.append(Spacer(1, 0.3 * cm))

    pares = [
        ("Estudiante", nombre),
        ("Código", codigo),
        ("Carrera", carrera),
        ("Ciclo", ciclo),
    ]
    elems.append(Paragraph("Datos del estudiante", h2))
    elems.append(_tabla_pares(pares, normal))

    elems.append(Paragraph("Diagnóstico académico", h2))
    if diagnostico:
        horas, promedio, _, estres, motivacion, procrast, puntaje, riesgo, fecha = diagnostico
        pares_diag = [
            ("Horas de estudio por día", horas),
            ("Promedio ponderado", promedio),
            ("Puntaje de riesgo", f"{puntaje}/100"),
            ("Nivel de riesgo", riesgo),
            ("Estrés", estres),
            ("Motivación", motivacion),
            ("Procrastinación", procrast),
            ("Fecha", fecha),
        ]
        if diagnostico_detallado:
            pares_diag.extend([
                ("Estado de ánimo", diagnostico_detallado.get("indice_estado_animo")),
                ("Alerta emocional", "Sí" if diagnostico_detallado.get("alerta_emocional") == 1 else "No"),
            ])
        elems.append(_tabla_pares(pares_diag, normal))

        if diagnostico_detallado:
            textos_ia = [
                ("Diagnóstico general IA", diagnostico_detallado.get("diagnostico_general_ia", "-")),
                ("Recomendación para el estudiante", diagnostico_detallado.get("recomendacion_estudiante_ia", "-")),
                ("Recomendación para tutoría", diagnostico_detallado.get("recomendacion_tutoria_ia", "-")),
            ]
            for titulo, texto in textos_ia:
                elems.append(Paragraph(titulo, h2))
                elems.append(_paragraph(texto, normal))
                elems.append(Spacer(1, 0.15 * cm))
    else:
        elems.append(Paragraph("No hay diagnóstico registrado.", normal))

    elems.append(Paragraph("Resumen de tareas", h2))
    pares_tareas = [
        ("Total de tareas", resumen_tareas.get("total", 0)),
        ("Completadas", resumen_tareas.get("completadas", 0)),
        ("Pendientes", resumen_tareas.get("pendientes", 0)),
        ("Alta prioridad", resumen_tareas.get("alta_prioridad", 0)),
        ("Cumplimiento", f"{resumen_tareas.get('porcentaje_cumplimiento', 0)}%"),
    ]
    elems.append(_tabla_pares(pares_tareas, normal))

    elems.append(Paragraph("Cursos registrados", h2))
    if cursos:
        data = [[_paragraph("Curso", small), _paragraph("Docente", small), _paragraph("Créditos", small), _paragraph("Dif.", small), _paragraph("Estado", small)]]
        for c in cursos:
            data.append([_paragraph(c[1], small), _paragraph(c[2], small), _paragraph(c[3], small), _paragraph(c[4], small), _paragraph(c[5], small)])
        tabla = Table(data, colWidths=[4.5 * cm, 4.0 * cm, 1.7 * cm, 1.3 * cm, 3.0 * cm], repeatRows=1)
        tabla.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(tabla)
    else:
        elems.append(Paragraph("No hay cursos registrados.", normal))

    elems.append(PageBreak())
    elems.append(Paragraph("Tareas registradas", h2))
    if tareas:
        data = [[_paragraph("Tipo", small), _paragraph("Actividad", small), _paragraph("Curso", small), _paragraph("Entrega", small), _paragraph("Estado", small)]]
        for t in tareas:
            # Formato actual: ID, Tipo, Actividad, Curso, Fecha, Prioridad, Estado.
            if len(t) >= 7:
                data.append([_paragraph(t[1], small), _paragraph(t[2], small), _paragraph(t[3], small), _paragraph(t[4], small), _paragraph(t[6], small)])
            else:
                data.append([_paragraph("Tarea", small), _paragraph(t[1], small), _paragraph(t[2], small), _paragraph(t[3], small), _paragraph(t[5], small)])
        tabla = Table(data, colWidths=[2.6 * cm, 4.2 * cm, 4.0 * cm, 2.3 * cm, 2.3 * cm], repeatRows=1)
        tabla.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(tabla)
    else:
        elems.append(Paragraph("No hay tareas registradas.", normal))

    doc.build(elems)
    output.seek(0)
    return output.getvalue()
